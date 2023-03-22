# from selenium import webdriver
# import csv
# import openpyxl
# import time
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException
# options = webdriver.ChromeOptions()
# # Read in list of proxies from CSV file
# proxies = []
# with open('active_proxies.csv', 'r') as f:
#     reader = csv.reader(f)
#     for row in reader:
#         proxies.append(row[0])
# baseurl = 'https://www.inmyarea.com/utilities?zipcode='
# urls = []
#
# wb = openpyxl.load_workbook('US.xlsx')
# ws = wb.active
# cloumn = ws['A']
# for cell in cloumn:
#     c = f'{cell.value}'
#     urls.append(baseurl + c)
#
# zipcode = []
#
#
# for i in range(0, len(proxies)):
#     try:
#         print("Proxy selected: {}".format(proxies[i]))
#         options = webdriver.ChromeOptions()
#         options.add_argument('--proxy-server={}'.format(proxies[i]))
#         driver = webdriver.Chrome(options=options, executable_path=r'G:/chromedriver.exe')
#         main_url = urls[1:101]
#         for data in main_url:
#             try:
#                 print("Proxy selected: {}".format(proxies[i]))
#
#                 options.add_argument('--proxy-server={}'.format(proxies[i]))
#                 driver = webdriver.Chrome(options=options, executable_path=r'G:/chromedriver.exe')
#
#                 driver.get(data)
#                 time.sleep(1)
#                 proviers = driver.find_elements(By.CSS_SELECTOR, '.provider-name a')
#                 try:
#                     type = driver.find_elements(By.CSS_SELECTOR,
#                                                 '#content-left > section > table > tbody > tr:nth-child(1) > td:nth-child(3)')
#                 except:
#                     type = None
#                 try:
#                     contact = driver.find_elements(By.CSS_SELECTOR,
#                                                    '#content-left > section > table > tbody > tr:nth-child(1) > td.text-right.provider-phone.hidden-xs > a.phone.swappable')
#                 except:
#                     contact = None
#                 try:
#                     link = driver.find_elements(By.XPATH,
#                                                 '//*[@id="content-left"]/section/table/tbody/tr[1]/td[2]/a//*[@href]')
#                 except:
#                     link = None
#                 temporary_data = {
#                     'provider': proviers[i].text,
#                     'type': type[i].text,
#                     'contact': contact[i].text,
#                     'link': link
#                 }
#                 zipcode.append(temporary_data)
#
#             except Exception:
#                 driver.quit()
from selenium import webdriver
import csv
import openpyxl
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

options = webdriver.ChromeOptions()
# Read in list of proxies from CSV file
proxies = []
urls = []
zipcode = []
baseurl = "https://www.inmyarea.com/utilities?zipcode="

with open("active_proxies.csv", "r") as f:
    reader = csv.reader(f)
    for row in reader:
        proxies.append(row[0])

wb = openpyxl.load_workbook("US.xlsx")
ws = wb.active
cloumn = ws["A"]
for cell in cloumn:
    c = f"{cell.value}"
    urls.append(baseurl + c)
main_url = urls[1:1000]
for i in range(1, len(proxies)):
    print(f"Proxy selected: {proxies[i]}")
    options = webdriver.ChromeOptions()
    options.add_argument(f"--proxy-server={proxies[i]}")
    driver = webdriver.Chrome(options=options, executable_path=r"G:/chromedriver.exe")

    for data in main_url:
        try:
            print("Proxy selected: {}".format(proxies[i]))

            options.add_argument('--proxy-server={}'.format(proxies[i]))
            driver = webdriver.Chrome(options=options, executable_path=r'G:/chromedriver.exe')

            driver.get(data)
            time.sleep(1)

            # providers = driver.find_elements(By.CSS_SELECTOR, '#content-left > section > table > tbody > tr:nth-child(1) > td:nth-child(3)')
            providers = driver.find_elements(By.CSS_SELECTOR, '.hidden-xs+ td')
            name = driver.find_elements(By.CSS_SELECTOR,'.provider-name a')



            for i in range(len(providers)):
                temporary_data = {
                    'provider': providers[i].text,

                }
                zipcode.append(temporary_data)

        except Exception:
            driver.quit()
    break

import pandas
df = pandas.DataFrame(zipcode)
df.to_csv('zipcodec.csv')
