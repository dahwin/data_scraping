# import openpyxl
# import pandas as pd
# import time
# from selenium.webdriver.common.keys import Keys
# from openpyxl import load_workbook
# from selenium.webdriver.common.by import By
# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
#
# options = Options()
# options.add_argument('--disable-blink-features=AutomationControlled')
# options.headless = False
# driver = webdriver.Chrome(options=options)
# from csv import reader
# baseurl = 'https://www.inmyarea.com/utilities?zipcode='
# urls = []
#
#
# wb = openpyxl.load_workbook('US.xlsx')
# ws = wb.active
# cloumn = ws['A']
# for cell in cloumn:
#     c = f'{cell.value}'
#     urls.append(baseurl+c)
#
# zipcode = []
# main_url = urls[1:101]
# for data in  main_url:
#     driver.get(data)
#     time.sleep(5)
#
#     proviers = driver.find_elements(By.CSS_SELECTOR,'.provider-name a')
#     try:
#       type = driver.find_elements(By.CSS_SELECTOR,'#content-left > section > table > tbody > tr:nth-child(1) > td:nth-child(3)')
#     except:
#       type = None
#     try:
#        contact = driver.find_elements(By.CSS_SELECTOR,'#content-left > section > table > tbody > tr:nth-child(1) > td.text-right.provider-phone.hidden-xs > a.phone.swappable')
#     except:
#         contact = None
#     try:
#
#        link = driver.find_elements(By.XPATH,'//*[@id="content-left"]/section/table/tbody/tr[1]/td[2]/a//*[@href]')
#     except:
#         link = None
#
#     for i in range(len(proviers)):
#         temporary_data = {
#             'provider': proviers[i].text,
#             'type': type[i].text,
#             'contact':contact[i].text,
#             'link': link
#         }
#         zipcode.append(temporary_data)

import openpyxl
import pandas as pd
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

options = Options()
options.add_argument('--disable-blink-features=AutomationControlled')
options.headless = False
driver = webdriver.Chrome(options=options)
from csv import reader

baseurl = 'https://www.inmyarea.com/utilities?zipcode='
urls = []

wb = openpyxl.load_workbook('US.xlsx')
ws = wb.active
cloumn = ws['A']
for cell in cloumn:
    c = f'{cell.value}'
    urls.append(baseurl + c)

zipcode = []
main_url = urls[1:101]
for data in main_url:
    driver.get(data)
    time.sleep(5)

    providers = driver.find_elements(By.CSS_SELECTOR, '.provider-name a')
    types = driver.find_elements(By.CSS_SELECTOR,
                                 '#content-left > section > table > tbody > tr:nth-child(1) > td:nth-child(3)')
    contacts = driver.find_elements(By.CSS_SELECTOR,
                                    '#content-left > section > table > tbody > tr:nth-child(1) > td.text-right.provider-phone.hidden-xs > a.phone.swappable')
    links = driver.find_elements(By.XPATH, '//*[@id="content-left"]/section/table/tbody/tr[1]/td[2]/a//*[@href]')

    for i in range(min(len(providers), len(types), len(contacts), len(links))):
        temporary_data = {
            'provider': providers[i].text,
            'type': types[i].text,
            'contact': contacts[i].text,
            'link': links[i].get_attribute('href')
        }
        zipcode.append(temporary_data)

